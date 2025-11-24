from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl import load_workbook

# Cores usadas para Status
STATUS_COLORS = {
    "Viabilidade Expressa": "90EE90",
    "Dentro": "90EE90",
    "Próximo à mancha": "ffeb9c",
    "Rede PTP": "ffeb9c",
    "Inviável": "FF7F7F",
    "Coordenada Inválida": "D3D3D3"
}

# Colunas adicionadas pela análise
NOVAS_COLUNAS_ANALISE = [
    "Status",
    "Mancha GPON",
    "Dist. GPON (mts)",
    "Rede PTP"
]

# Cor do cabeçalho para colunas novas
HEADER_NOVO = "4472C4"  # azul escuro profissional

def autoajuste(path_excel: str):
    wb = load_workbook(path_excel)
    ws = wb.active

    # Congelar cabeçalho
    ws.freeze_panes = "A2"

    # ↓↓↓ Identificar colunas novas ↓↓↓
    col_novas = {}
    for col in range(1, ws.max_column + 1):
        header = ws.cell(row=1, column=col).value
        if header in NOVAS_COLUNAS_ANALISE:
            col_novas[header] = col

    col_status = col_novas.get("Status")

    # =====================
    #  1. COLORIR LINHAS
    # =====================
    if col_status:
        for row in range(2, ws.max_row + 1):
            status = ws.cell(row=row, column=col_status).value
            cor = STATUS_COLORS.get(status)
            if cor:
                fill = PatternFill(start_color=cor, end_color=cor, fill_type="solid")
                for header, col in col_novas.items():
                    ws.cell(row=row, column=col).fill = fill

    # =============================================
    #  2. APLICAR BORDA EM TODA A PLANILHA
    # =============================================
    thin = Side(border_style="thin", color="000000")
    border = Border(top=thin, left=thin, right=thin, bottom=thin)

    for row in ws.iter_rows():
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="center")

    # ============================================================
    #  3. COLORIR CABEÇALHO DAS NOVAS COLUNAS
    # ============================================================
    header_fill = PatternFill(start_color=HEADER_NOVO, end_color=HEADER_NOVO, fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    for header, col in col_novas.items():
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font

    # =============================================
    #  4. AUTOAJUSTE DE COLUNA
    # =============================================
    for column_cells in ws.columns:
        max_length = 0
        column_letter = column_cells[0].column_letter

        for cell in column_cells:
            try:
                text = str(cell.value)
                max_length = max(max_length, len(text))
            except:
                pass

        ws.column_dimensions[column_letter].width = max_length + 2

    wb.save(path_excel)
